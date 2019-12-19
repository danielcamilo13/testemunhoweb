# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.http import HttpResponse,HttpResponseRedirect
from django.shortcuts import render
import os, sys,openpyxl,re,calendar
from datetime import datetime
from .models import irmaos,dias,designacao
import json, csv
from openpyxl.styles import fills,PatternFill,Border, Side, Alignment, Protection, Font, Color, colors
from openpyxl import Workbook
from .tratamento import tratamento,spreadsheet_reader
from django.core.files.storage import FileSystemStorage
from django.core.serializers import serialize

def index(request):
    return 'index'

def designar(request):
    qs={}
    hoje =datetime.now()
    mes = hoje.strftime('%m')
    ano = hoje.strftime('%Y')
    qs = request.META['QUERY_STRING']
    return render(request,'cadastro/designar.html',{'qs':qs,'hoje':hoje,'mes':mes,'ano':ano})


def gerador(request):
    BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    BASE_DIR = os.path.join(BASE,'logs')
    hoje = datetime.now()
    d = hoje.strftime('%d')
    qs = str(request.META['QUERY_STRING'])
    f = open(os.path.join(BASE_DIR,'lista_mes.json'),'w')
    list_mes = {}
    arqmes = {'campo':'nome'}
    mensagem = 'Mensagem de teste'
    external = sys.stdout
    sys.stdout = open(os.path.join(BASE_DIR,'backlog.log'),'w')
    arqmes = open(os.path.join(BASE_DIR,'mes.txt'),'w')
    lista_mes = []
    wbMes = Workbook()
    wsMes = wbMes.active
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'))
    fill = PatternFill("solid", fgColor="DDDDDD")
    mes_json = open(os.path.join(BASE_DIR,'file.json'),'w')

    if 'ano' in request.POST:
        # a = str(request.POST['ano'])
        mes = int(request.POST['mes'])
        mes_label = calendar.month_name[mes]
        # mes_label = mes.strftime('%B')
        planMes = os.path.join(BASE_DIR, mes_label) + '.xlsx'
        # mes_ext = datetime.strptime(m,'%m')
        mensagem = mes_label
        ultimo_dia = calendar.monthrange(int(d), int(mes))[1]
        campos = [f.name for f in dias._meta.get_fields()][3:]
        expurgo_dia = []
        designado_dia = []
        lista_segunda = {}; lista_terca = {};lista_quarta ={};lista_quinta ={}; lista_sexta ={};lista_sabado ={}
        for dia_mes in range(1,ultimo_dia+1):
            dia_semana = calendar.weekday(int(d), int(mes), dia_mes)
            if dia_semana==0:
                ds='Segunda-feira'
                dia=filling_header()
                dia['dia_mes']=dia_mes
                dia['dia_semana']=ds
                par = dia_mes % 2
                lista_dia = lista_segunda # esta linha e usada para zerar a variavel lista_dia pois esta e usada para preenchimento do dia final
                filled_day = filled_each_day(ds,dia,par,dia_mes,lista_dia,expurgo_dia,designado_dia)
                lista_segunda = filled_day
                lista_dia = lista_segunda
                # lista_dia=sorted(lista_segunda.values()); print('lista segunda %s'%str(lista_dia))
                # print('Periodos em branco no dia %s %s: %s'%(str(dia_mes),ds,lista_dia.count('EMPTY')))
                lista_dia = doublecheck(lista_segunda.items(),ds,designado_dia)
                lista_mes.append(sorted(lista_dia.items()))
                swap_file(arqmes,lista_dia)
                # swap_file(arqmes,filled_day)
                swap_xls(lista_dia,wsMes,dia_mes)
                grava_json(BASE_DIR,lista_mes)
                # rec_dict(BASE_DIR, lista_mes)
                list_mes[ds]='culinaria'
                for k,v in lista_segunda.items():
                    # list_mes[lista_segunda.keys()]=lista_segunda.values()
                    print(' TESTANDO DICIONARIO %s %s'%(k,v))
                json.dump(lista_segunda,mes_json,indent=4)
                # print(json.loads(json_data))

            elif dia_semana==1:
                ds='Terca-feira';print(ds)
                dia=filling_header()
                dia['dia_mes']=dia_mes
                dia['dia_semana']=ds
                par = dia_mes % 2
                lista_dia = lista_terca
                filled_day = filled_each_day(ds,dia,par,dia_mes,lista_dia,expurgo_dia,designado_dia)
                lista_terca = filled_day
                lista_dia = lista_terca
                # lista_dia=sorted(lista_terca.values()); print('lista Terca %s'%lista_dia)
                # print('Periodos em branco no dia %s %s: %s'%(str(dia_mes),ds,lista_dia.count('EMPTY')))
                lista_dia = doublecheck(lista_terca.items(),ds,designado_dia)
                lista_mes.append(sorted(lista_dia.items()))
                swap_file(arqmes,lista_dia)
                # swap_file(arqmes,filled_day)
                swap_xls(lista_dia,wsMes,dia_mes)
                grava_json(BASE_DIR,lista_mes)
                # rec_dict(BASE_DIR, lista_mes)
                json.dump(lista_terca,mes_json,indent=4)
                # print(json.loads(json_data))

            elif dia_semana==2:
                ds='Quarta-feira';print(ds)
                dia=filling_header()
                dia['dia_mes']=dia_mes
                dia['dia_semana']=ds
                par = dia_mes % 2
                lista_dia = lista_quarta
                filled_day = filled_each_day(ds,dia,par,dia_mes,lista_dia,expurgo_dia,designado_dia)
                lista_quarta = filled_day
                lista_dia = lista_quarta
                # lista_dia= sorted(lista_quarta.values()); print('lista quarta %s'%lista_dia)
                # print('Periodos em branco no dia %s %s: %s'%(str(dia_mes),ds,lista_dia.count('EMPTY')))
                lista_dia = doublecheck(lista_quarta.items(),ds,designado_dia)
                lista_mes.append(sorted(lista_dia.items()))
                swap_file(arqmes, lista_dia)
                # swap_file(arqmes,filled_day)
                swap_xls(lista_dia,wsMes,dia_mes)
                grava_json(BASE_DIR,lista_mes)
                # rec_dict(BASE_DIR, lista_mes)
                json_data = json.dump(lista_quarta,mes_json,indent=4)
                # print(json.loads(json_data))

            elif dia_semana==3:
                ds='Quinta-feira';print(ds)
                dia=filling_header()
                dia['dia_mes']=dia_mes
                dia['dia_semana']=ds
                par = dia_mes % 2
                lista_dia = lista_quinta
                filled_day = filled_each_day(ds,dia,par,dia_mes,lista_dia,expurgo_dia,designado_dia)
                lista_quinta = filled_day
                lista_dia = lista_quinta
                # lista_dia= sorted(lista_quinta.values()); print('lista quinta %s'%lista_dia)
                # print('Periodos em branco no dia %s %s: %s'%(str(dia_mes),ds,lista_dia.count('EMPTY')))
                # lista_dia = doublecheck(lista_quinta.items(),ds,designado_dia)
                lista_mes.append(sorted(lista_dia.items()))
                swap_file(arqmes, lista_dia)
                # swap_file(arqmes,filled_day)
                swap_xls(lista_dia,wsMes,dia_mes)
                grava_json(BASE_DIR,lista_mes)
                # rec_dict(BASE_DIR, lista_mes)
                json_data = json.dump(lista_quinta,mes_json,indent=4)
                # print(json.loads(json_data))

            elif dia_semana==4:
                ds='Sexta-feira';print(ds)
                dia=filling_header()
                dia['dia_mes']=dia_mes
                dia['dia_semana']=ds
                par = dia_mes % 2
                lista_dia=lista_sexta
                filled_day = filled_each_day(ds,dia,par,dia_mes,lista_dia,expurgo_dia,designado_dia)
                lista_sexta = filled_day
                lista_dia=lista_sexta
                # lista_dia = doublecheck(lista_sexta.items(),ds,designado_dia)
                lista_mes.append(sorted(lista_dia.items()))
                swap_file(arqmes, lista_dia)
                # swap_file(arqmes,filled_day)
                swap_xls(lista_dia,wsMes,dia_mes)
                grava_json(BASE_DIR,lista_mes)
                # rec_dict(BASE_DIR, lista_mes)
                json_data = json.dump(lista_sexta,mes_json,indent=4)
                # print(json.loads(json_data))

            elif dia_semana==5:
                ds='Sabado';print(ds)
                dia=filling_header()
                dia['dia_mes']=dia_mes
                dia['dia_semana']=ds
                par = dia_mes % 2
                lista_dia = lista_sabado
                filled_day = filled_each_day(ds,dia,par,dia_mes,lista_dia,expurgo_dia,designado_dia)
                lista_dia=lista_sabado
                lista_sabado = filled_day
                lista_dia = lista_sabado
                # lista_dia= sorted(lista_sabado.values())
                print('lista sabado %s'%lista_dia)
                # print('Periodos em branco no dia %s %s: %s'%(str(dia_mes),ds,lista_dia.count('EMPTY')))
                lista_dia = doublecheck(lista_sabado.items(),ds,designado_dia)
                lista_mes.append(sorted(lista_dia.items()))
                swap_file(arqmes,lista_dia)
                swap_xls(lista_dia,wsMes,dia_mes)
                grava_json(BASE_DIR,lista_mes)
                # rec_dict(BASE_DIR, lista_mes)
                json_data = json.dump(lista_sabado,mes_json,indent=4)
                # print(json.loads(json_data))
    mes_json = open(os.path.join(BASE_DIR,'file.json'),'r')
    contagem_irmaos(lista_mes)
    arqmes.close()
    sys.stdout.close()
    sys.stdout = external
    arqmes = open(os.path.join(BASE_DIR, 'mes.txt'), 'r')
    ultima_linha=wsMes.max_row+2
    rowdate = datetime.today()

    wsMes.merge_cells(start_row=ultima_linha+2,start_column=1,end_row=ultima_linha+2,end_column=2)
    wsMes.cell(row=ultima_linha,column=1).value='Data de geracao'
    wsMes.merge_cells(start_row=ultima_linha+2,start_column=3,end_row=ultima_linha+2,end_column=6)
    wsMes.cell(row=ultima_linha,column=3).value=rowdate
    wsMes.insert_rows(1,1)
    wsMes.merge_cells(start_row=1,start_column=1,end_row=1,end_column=3)
    wsMes.cell(row=1,column=1).value=mes_label
    wsMes.insert_rows(2,1)
    wsMes.cell(row=2,column=1).value='Dia Mes'
    wsMes.cell(row=2,column=2).value='Dia Semana'
    wsMes.merge_cells(start_row=2,start_column=3,end_row=2,end_column=5)
    wsMes.cell(row=2,column=3).value='Periodo 1'
    wsMes.merge_cells(start_row=2,start_column=6,end_row=2,end_column=7)
    wsMes.cell(row=2,column=6).value='Periodo 2'
    wsMes.merge_cells(start_row=2,start_column=8,end_row=2,end_column=9)
    wsMes.cell(row=2,column=8).value='Periodo 3'
    wsMes.merge_cells(start_row=2,start_column=10,end_row=2,end_column=11)
    wsMes.cell(row=2,column=10).value='Periodo 4'
    wsMes.merge_cells(start_row=2,start_column=12,end_row=2,end_column=13)
    wsMes.cell(row=2,column=12).value='Periodo 5'
    fmtPlan(wsMes)
    wbMes.save(filename=planMes)
    return render(request,'cadastro/resultado.html',{'mensagem':mensagem,'lista_mes':lista_mes,'qs':qs,'hoje':hoje,'mes_label':mes_label})

def localiza_dias(ds,dia_mes):
    dia = filling_header()
    dia['dia_mes'] = dia_mes
    dia['dia_semana'] = ds
    par = dia_mes % 2
    return dia,par

def swap_file(arqmes,dia):
    print('preenchido com os valores:  %s '%(dia))
    dia_ordenado = sorted(dia.items())
    arqmes.write(str(dia_ordenado)+'\n')

def swap_xls(dia,wsMes,dia_mes):
    wsMes.cell(row=dia_mes,column=1).value=dia['dia_mes']
    wsMes.cell(row=dia_mes,column=2).value=dia['dia_semana']
    wsMes.cell(row=dia_mes,column=3).value=dia['p1']
    wsMes.cell(row=dia_mes,column=4).value=dia['p1_1']
    wsMes.cell(row=dia_mes,column=5).value=dia['p1_2']
    wsMes.cell(row=dia_mes,column=6).value=dia['p2']
    wsMes.cell(row=dia_mes,column=7).value=dia['p2_1']
    wsMes.cell(row=dia_mes,column=8).value=dia['p3']
    wsMes.cell(row=dia_mes,column=9).value=dia['p3_1']
    wsMes.cell(row=dia_mes,column=10).value=dia['p4']
    wsMes.cell(row=dia_mes,column=11).value=dia['p4_1']
    wsMes.cell(row=dia_mes,column=12).value=dia['p5']
    wsMes.cell(row=dia_mes,column=13).value=dia['p5_1']

def filling_header():
    line = {'dia_mes':'EMPTY',
            'dia_semana':'EMPTY',
            'p1':'EMPTY','p1_1':'EMPTY','p1_2':'EMPTY',
            'p2':'EMPTY','p2_1':'EMPTY',
            'p3':'EMPTY','p3_1':'EMPTY',
            'p4':'EMPTY','p4_1':'EMPTY',
            'p5':'EMPTY','p5_1':'EMPTY'}
    return line

def filled_each_day(ds,dia,par,dia_mes,lista_dia,expurgo_dia,designado_dia):
    if par == 0:
        print('Dia %s eh par' % dia_mes)
        adv = 'Par'
    else:
        print('Dia %s eh impar' % dia_mes)
        adv = 'Impar'

    contagem = [i['irmao__nm'] for i in consulta_irmaos() if i['dia_semana']==ds if i['adv']==adv or i['adv']=='Nulo']
    print('%s'%'*'*30)
    print('Elegiveis para %s. contagem %s'%(ds,len(contagem)))
    for cont in contagem:
        print(' *** irmaos %s ***'%cont)
    print('%s'%'*'*30)

    periodo1=[]; periodo2=[]; periodo3=[];periodo4=[];periodo5=[]; expurgo=[]
    periodo1_gr=[]; periodo2_gr=[]; periodo3_gr=[];periodo4_gr=[];periodo5_gr=[]
    '''
    AQUI NESTE TRECHO EXCECAO ADICIONAL QUE TRABALHARA COM PECULIARIDADES QUE PREENCHERAO AS PREFERENCIAS
    '''
    #antes do tratamento recuperar todos os campos de periodos
    fill_header = filling_header()
    trat = tratamento(dia,dia_mes,ds,par,fill_header,designado_dia)
    for i in consulta_irmaos():
        p1=i['p1']; p1_1=i['p1_1']; p1_2=i['p1_2']; p2  =i['p2']; p2_1=i['p2_1']; p3  =i['p3']; p3_1=i['p3_1']; p4  =i['p4']; p4_1=i['p4_1']; p5  =i['p5']; p5_1=i['p5_1']
        irmao = i['irmao__nm']
        conjuge = i['irmao__conjuge']
        genero = i['irmao__gr']
        maximo = i['irmao__maximo']
        pref = i['preferencial']
        if i['dia_semana'] == ds:
            cl_day = cleaner_days(ds,dia)
            if pref == True:
                print('O irmao %s tem preferencias neste dia '%irmao)
            if not re.search(irmao,str(lista_dia)):
                print('Irmaos PREVIAMENTE cadastrados %s'%lista_dia)
                for key,value in sorted(dia.items()):
                    if value=='EMPTY':
                        print('--> Chave em branco %s'%key)
                        if i[key]==True:
                            if i['adv'] == adv or i['adv'] == 'Nulo':
                                print('Varredura do irmao(a) %s' % irmao)
                                print('Irmao %s disponivel para trabalho no PERIODO %s e trabalha em dias %ses'%(irmao,key,adv))
                                #preenchendo periodos
                                if key=='p1' or key=='p1_1' or key=='p1_2':
                                    periodo = periodo1
                                    periodo_gr = periodo1_gr
                                    p1 = 'p1'
                                    p2 = 'p1_2'
                                    preenchimento = dados_preenchimento(periodo, periodo_gr, irmao, conjuge, genero, dia, ds, key,
                                                                        p1, p2,i,maximo,lista_dia,expurgo_dia,designado_dia)
                                elif key == 'p2' or key == 'p2_1':
                                    periodo = periodo2
                                    periodo_gr = periodo2_gr
                                    p1 = 'p2'
                                    p2 = 'p2_1'
                                    preenchimento = dados_preenchimento(periodo, periodo_gr, irmao, conjuge, genero, dia,ds, key,
                                                                        p1, p2, i,maximo,lista_dia,expurgo_dia,designado_dia)
                                elif key == 'p3' or key == 'p3_1':
                                    periodo = periodo3
                                    periodo_gr = periodo3_gr
                                    p1 = 'p3'
                                    p2 = 'p3_1'
                                    preenchimento = dados_preenchimento(periodo, periodo_gr, irmao, conjuge, genero, dia,ds, key,
                                                                        p1, p2, i,maximo,lista_dia,expurgo_dia,designado_dia)
                                elif key == 'p4' or key == 'p4_1':
                                    periodo = periodo4
                                    periodo_gr = periodo4_gr
                                    p1 = 'p4'
                                    p2 = 'p4_1'
                                    preenchimento = dados_preenchimento(periodo, periodo_gr, irmao, conjuge, genero, dia,ds, key,
                                                                        p1, p2, i,maximo,lista_dia,expurgo_dia,designado_dia)
                                elif key == 'p5' or key == 'p5_1':
                                    periodo = periodo5
                                    periodo_gr = periodo5_gr
                                    p1 = 'p5'
                                    p2 = 'p5_1'
                                    preenchimento = dados_preenchimento(periodo, periodo_gr, irmao, conjuge, genero, dia,ds, key,
                                                                    p1, p2, i,maximo,lista_dia,expurgo_dia,designado_dia)
            else:
                print('Irmao(a) %s ja foi PREVIAMENTE designado' %irmao)
    print('%s \n Dia da semana preenchido %s : %s \n%s' % ('*'*50,ds,dia,'*'*50))
    lista_dia = dia.values()
    return dia#,lista_dia

def consulta_irmaos():
    orm_dia = dias.objects.select_related('irmao').values('dia_semana','irmao','irmao__nm','p1','p2','p3','p4','p5','p1_1','p1_2','p2_1','p3_1','p4_1','p5_1','irmao__habilitado','irmao__estado_civil','irmao__conjuge','irmao__gr','adv','irmao__maximo','preferencial','irmao__trio','irmao__privilegio','irmao__dianteira','irmao__excecao_dia','irmao__excecao_nome').filter(irmao__habilitado='True').order_by('?')
    return orm_dia

def showing(request):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    with open(os.path.join(BASE_DIR, 'data.json')) as datajson:
        data = json.load(datajson)
        # datajson = open(os.path.join(BASE_DIR,'data.json'))
        # datajson = {'foo': 'bar', 'hello': 'world'}
    return HttpResponse(json.load(data,encoding=None,),content_type='application/json')

def fmtPlan(wsMes):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill = PatternFill("solid", fgColor="DDDDDD")
    for row in range(1,wsMes.max_row+1):
        for col in range(1,wsMes.max_column+1):
            wsMes.cell(row=row,column=col).border=thin_border
            if row==1:
                wsMes.cell(row=row,column=col).font=Font(name='Times New Roman',size='14',bold=True,italic=False)
            if row==2:
                wsMes.cell(row=row,column=col).font=Font(name='Times New Roman',size='12',bold=False,italic=True)
                wsMes.cell(row=row,column=col).fill = openpyxl.styles.PatternFill('solid', openpyxl.styles.colors.YELLOW)
            if col<=2 and row>2:
                wsMes.cell(row=row,column=col).font=Font(name='Times New Roman',size='12',bold=True,italic=False,color=colors.BLUE)

def conjuge_dados(dia,irmao,ds,key,conjuge):
    casado = [{'irmao':i['irmao__nm'],'p1':i['p1'],'p1_1':i['p1_1'],'p1_2':i['p1_2'],'p2':i['p2'],'p2_1':i['p2_1'],'p3':i['p3'],'p3_1':i['p3_1'],'p4':i['p4'],'p4_1':i['p4_1'],'p5':i['p5'],'p5_1':i['p5_1']} for i in consulta_irmaos() if conjuge == i['irmao__nm'] and ds==i['dia_semana']]
    return casado

def doublecheck(lista_dia,ds,designado_dia):
    lista_dia = dict(lista_dia)
    if 'EMPTY' in lista_dia.values(): #verifica se no dia existe um campo vazio
        gender =''
        for k, v in lista_dia.items():
            if v == 'EMPTY':
                print('SEGUNDA RODADA %s'%(k))
                if k =='p1':
                    print('parceiro: '+lista_dia['p1_1'])
                    gender = [genero['irmao__nm']+genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p1_1']]
                    print('Genero %s'%str(gender))
                elif k =='p1_1':
                    print('parceiro: '+lista_dia['p1'])
                    gender = [genero['irmao__nm'] + genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p1']]
                    print('Genero %s'%str(gender))
                elif k =='p2':
                    print('parceiro: '+lista_dia['p2_1'])
                    gender = [genero['irmao__nm'] + genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p2_1']]
                    print('Genero %s'%str(gender))
                elif k =='p2_1':
                    print('parceiro: '+lista_dia['p2'])
                    gender = [genero['irmao__nm'] + genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p2']]
                    print('Genero %s'%str(gender))
                elif k =='p3':
                    print('parceiro: '+lista_dia['p3_1'])
                    gender = [genero['irmao__nm']+ genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p3_1']]
                    print('Genero %s'%str(gender))
                elif k =='p3_1':
                    print('parceiro: '+lista_dia['p3'])
                    gender = [genero['irmao__nm']+ genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p3']]
                    print('Genero %s'%str(gender))
                elif k =='p4':
                    print('parceiro: '+lista_dia['p4_1'])
                    gender = [genero['irmao__nm']+genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p4_1']]
                    print('Genero %s'%str(gender))
                elif k =='p4_1':
                    print('parceiro: '+lista_dia['p4'])
                    gender = [genero['irmao__nm']+genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p4']]
                    print('Genero %s'%str(gender))
                elif k =='p5':
                    print('parceiro: '+lista_dia['p5_1'])
                    gender = [genero['irmao__nm']+ genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p5_1']]
                    print('Genero %s'%str(gender))
                elif k =='p5_1':
                    print('parceiro: '+lista_dia['p5'])
                    gender = [genero['irmao__nm']+genero['irmao__gr'] for genero in consulta_irmaos() if genero['irmao__nm'] == lista_dia['p5']]
                    print('Genero %s'%str(gender))

                for ir in reversed(consulta_irmaos()):
                    if ir['dia_semana'] == ds:
                        if ir['irmao__nm'] not in lista_dia.values():
                            if ir[k]!=False:
                                if v =='EMPTY':
                                    if designado_dia.count(str(ir['irmao__nm'])) <= int(ir['irmao__maximo']):
                                        # if re.search(gender,ir['irmao__gr']):
                                        try:
                                            if re.search(r'%s',ir['irmao__gr'])%str(gender):
                                                    lista_dia[k] = ir['irmao__nm']
                                                    designado_dia.append(ir['irmao__nm'])
                                                    print('SOMA DEU CERTO AQUI %s %s %s'%(ir['irmao__nm'],designado_dia.count(str(ir['irmao__nm'])),ir['irmao__max']))
                                                    print('Irmao(a) %s foi ADICIONADO na segunda rodada %s = %s' % (ir['irmao__nm'], k,ir[k]))
                                        except:
                                            pass
                                        else:
                                            print('SEGUNDA RODADA NAO ROLOU')
                                    else:
                                        print('SOMA NAO DEU CERTO AQUI %s %s %s'%(ir['irmao__nm'],designado_dia.count(str(ir['irmao__nm'])),ir['irmao__maximo']))
    return lista_dia

def message():
    message1 = 'Mensagem de Teste 1'
    message2 = 'IRMAO habilitado:'
    message3 = 'GENERO no periodo '
    message4 = '**** Campo Par/Impar sao divergentes ***'
    message5 = 'Irmao nao foi incluido no periodo. Provalmente ja tenha sido designado. Segue periodo:'
    return message1,message2,message3,message4,message5

def dados_preenchimento(periodo,periodo_gr,irmao,conjuge,genero,dia,ds,key,p1,p2,i,maximo,lista_dia,expurgo_dia,designado_dia):
    print('Lista de Expurgo antes da analise %s'%sorted(expurgo_dia))
    if len(periodo)==0 or genero in periodo_gr: #lista de periodo preenchidos
        if int(maximo) > designado_dia.count(str(irmao)):
            if not irmao in dia.values():
                dia[key]=str(irmao)
                periodo_gr.append(str(genero))
                periodo.append(str(irmao))
                designado_dia.append(str(irmao))
                if genero=='masculino':
                    print('Irmao %s DESIGNADO na %s no periodo %s'%(irmao,ds,p1))
                else:
                    print('Irma %s DESIGNADA na %s no periodo %s' % (irmao, ds, p1))
                print('Genero %s adicionado no %s. %s'%(genero,key,periodo_gr))
                print('ADICIONADOS %s'%sorted(designado_dia))
                if conjuge is not None:
                    print('Irmao(a) %s e casado com %s' % (irmao, conjuge))
                    conjuge_periodo = conjuge_dados(dia, irmao, ds, key, conjuge)
                    for c in conjuge_periodo:
                        if c[p2] == True and dia[p2] == 'EMPTY':
                            print('Conjuge %s habilitado para ficar no mesmo horario (Periodo %s) %s' % (conjuge,p2,c[p2]))
                            dia[p2] = str(conjuge)
                            periodo.append(str(conjuge))
                            designado_dia.append(str(conjuge))
                        else:
                            print('Conjuge nao e voluntario para o mesmo horario')
                if irmao in expurgo_dia:
                    expurgo_dia.remove(irmao)
                    print('Irmao %s removido da lista de expurgo '%irmao)
            else:
                print('Irmao nao foi designado no periodo %s'%key)
    else:
        expurgo_dia.append(irmao)
        print('Irmao nao adicionado pois sao de generos diferentes')
        print('irmaos adicionado na lista de expurgo no preenchimento %s'%sorted(expurgo_dia))
    return dia

def cleaner_days(ds,dia):
    if ds == 'Terca-feira' or ds == 'Quarta-feira':
        dia['p5'] = ''; dia['p5_1'] = ''
    if ds == 'Terca-feira' or ds == 'Quarta-feira' or ds == 'Sexta-feira' or ds == 'Sabado' or ds == 'Segunda-feira':
        dia['p1_2'] = ''; dia['p3'] = ''; dia['p3_1'] = ''
    if ds == 'Sabado':
        dia['p1'] = ''; dia['p1_1'] = ''; dia['p1_2'] = ''; dia['p2'] = ''; dia['p2_1'] = ''; dia['p3'] = ''; dia['p3_1'] = ''
    if ds == 'Quinta-feira':
        dia['p2'] = ''; dia['p2_1'] = ''; dia['p1_2'] = ''; dia['p5'] = ''; dia['p5_1'] = ''; dia['p3'] = ''; dia['p3_1'] = ''; dia['p4'] = '';dia['p4_1'] = ''
    return ds,dia

def grava_json(BASE_DIR,lista_mes):
    with open(os.path.join(BASE_DIR,'lista_mes.json'),'a') as f:
        json.dump(lista_mes,f)

def contagem_irmaos(lista_mes):
    p = len(lista_mes)
    return str(p)

def querydict_to_dict(query_dict):
    data = {}
    for key in query_dict.keys():
        v = query_dict.getlist(key)
        data[key]=v
    return data

def confirmacao(request):
    BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    BASE_DIR = os.path.join(BASE,'logs')
    mes_json = open(os.path.join(BASE_DIR,'file_new.json'),'w')
    data = querydict_to_dict(request.POST)
    total_dias = len([v for v in request.POST.getlist('dia_mes')])
    myDictAll = []
    c = 1
    print('total de dias {}'.format(total_dias))
    while c <total_dias:
        myDict = {k:list(request.POST.getlist(k))[c] for k,v in request.POST.items() if k!='csrfmiddlewaretoken'}
        # print(myDict)
        gravar = designacao(mes=myDict['mes'],dia_mes=myDict['dia_mes'],dia_semana=myDict['dia_semana'],p1=myDict['p1'],p1_1=myDict['p1_1'],p1_2=myDict['p1_2'],p2=myDict['p2'],p2_1=myDict['p2_1'],p3=myDict['p3'],p3_1=myDict['p3_1'],p4=myDict['p4'],p4_1=myDict['p4_1'],p5=myDict['p5'],p5_1=myDict['p5_1'])
        myDictAll.append(myDict)
        gravar.save()
        c+=1
    return render(request,'cadastro/confirmacao.html',{'request':request,'data':data,'myDictAll':myDictAll})


def importar(request):
    context={}
    leitor={}
    fs = FileSystemStorage()
    if request.method=='POST':
        uploadFile=request.FILES['file_import']
        name = fs.save(uploadFile.name, uploadFile)
        leitor = spreadsheet_reader(uploadFile)
        context = leitor
    return render(request,'cadastro/importar.html',{'context':context})