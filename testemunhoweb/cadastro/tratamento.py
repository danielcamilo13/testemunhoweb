import os,sys,re
from .models import irmaos,dias
from openpyxl import load_workbook,Workbook
from django.core.files.storage import FileSystemStorage
import django_cal
from datetime import datetime

# O objetivo desta funcao é fazer o tratamento das excessoes preferenciais
def tratamento(dia,dia_mes,ds,par,fill_header,designado_dia):
    '''
    ==> Tratamento das seguintes excessoes
    Prioridade no dia
    Excecao para com dia
    Excecao para com nome - esta opcao AINDA ESTA DESABILITADA
    trio
    '''

    irmaos = consulta_irmaos_dia_semana(ds)
    print('%s Tratamento de excecoes %s'%('*'*20,'*'*20))
    print('{} - tratando o dia {} - {}'.format(datetime.now(),dia_mes,ds))
    print('{} - Tratamento - campos prenchidos do dicionario dia {}'.format(datetime.now(),dia))
    for irmao in irmaos:
        if irmao['preferencial']==True:
            if not re.search(str(dia_mes),str(irmao['irmao__excecao_dia'])):
                print('=====>Irmaos {} e preferencial na {}\n...e nao tem  excessao para trabalhar nas  {}'.format(irmao['irmao__nm'],irmao['dia_semana'],dia_mes))
                print('Nao foi detectado nenhuma excecao adicional considerando o dia  {}. Dias de excecoes  {}'.format(dia_mes,irmao['irmao__excecao_dia']))
                for k,v in sorted(irmao.items()):
                    if k in dia and irmao['irmao__nm'] not in designado_dia and v==True:
                        print('{} - ADICIONADO pois ele nao esta nas restricoes {}'.format(datetime.now(),irmao['irmao__nm']))
                        print('{} - CHAVE ADICIONADA {}'.format(datetime.now(),k))
                        designado_dia.append(irmao['irmao__nm'])
                        dia[k]=irmao['irmao__nm']

    return dia,designado_dia

def consulta_irmaos_dia_semana(ds):
    orm_dia = dias.objects.select_related('irmao').values('dia_semana','irmao','irmao__nm','p1','p2','p3','p4','p5','p1_1','p1_2','p2_1','p3_1','p4_1','p5_1','irmao__habilitado','irmao__estado_civil','irmao__conjuge','irmao__gr','adv','irmao__maximo','preferencial','irmao__trio','irmao__privilegio','irmao__dianteira','irmao__excecao_dia','irmao__excecao_nome').filter(irmao__habilitado='True',dia_semana=ds).order_by('?')
    return orm_dia

def consulta_irmao_dia():
    orm_dia = dias.objects.select_related('irmao').values('dia_semana','irmao','irmao__nm','p1','p2','p3','p4','p5','p1_1','p1_2','p2_1','p3_1','p4_1','p5_1','irmao__habilitado','irmao__estado_civil','irmao__conjuge','irmao__gr','adv','irmao__maximo','preferencial','irmao__trio','irmao__privilegio','irmao__dianteira','irmao__excecao_dia','irmao__excecao_nome').filter(irmao__habilitado='True').order_by('-dia_semana')
    return orm_dia

def consulta_irmaos():
    orm_dia = dias.objects.select_related('irmao').values('dia_semana','irmao','irmao__nm','p1','p2','p3','p4','p5','p1_1','p1_2','p2_1','p3_1','p4_1','p5_1','irmao__habilitado','irmao__estado_civil','irmao__conjuge','irmao__gr','adv','irmao__maximo','preferencial','irmao__trio','irmao__privilegio','irmao__dianteira','irmao__excecao_dia','irmao__excecao_nome').filter(irmao__habilitado='True').order_by('irmao__nm')
    return orm_dia

def spreadsheet_reader(spreadsheet):
    data = []
    dicionario = []
    wb = load_workbook(filename=spreadsheet)
    ws = wb.active
    # chaves = [ch for ch in range(1,ws.max_column)]
    # print('estas sao as chaves {}'.format(chaves))
    last_row = ws.max_row
    last_column = ws.max_column
    for r in range(1,last_row):
        lines = []
        for c in range(1,last_column):
            v = ws.cell(row=r,column=c).value
            if v is None:
                v=''
            lines.append(v)
        data.append(lines)
    return data

#funcao para tratar das excessoes dos irmaos
# Inicialmente está sendo tratado somente a excessão de dias
def excessoes(irmao,dia_mes):
    excp=''
    exc_dias = [exc for exc in irmaos.objects.values('excecao_dia').filter(nm=irmao)]
    for exc_item in exc_dias: # for para tirar o conteudo da lista
        for k,valores in exc_item.items(): #for para todos os valores que estao dentro do unico campo de excecao de dias
            if valores != None or valores != '':
                try:
                    for v in valores.split(','):
                        if int(v) == int(dia_mes):
                            print('dia com excecao LOCALIZADA {}***{}'.format(v,dia_mes))
                            excp=True
                except:
                    pass
    exc_nomes = [exc for exc in irmaos.objects.values('excecao_nome').filter(nm=irmao)]
    return excp,exc_nomes

def headers_dict():
    bla = 0
